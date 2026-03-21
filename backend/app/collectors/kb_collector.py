"""
KB 리브온 Excel 파일 수집기
파일: ☆ 1-1. KB 부동산 데이터 차트.xlsx (로컬 파일)
"""
import logging
from datetime import date, datetime
from pathlib import Path

import openpyxl
from app.config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()

KB_FILE = Path(settings.KB_FILE_PATH) if getattr(settings, "KB_FILE_PATH", None) else None

_SEARCH_PATHS = [
    Path(r"C:\Users\smk20\Desktop\1. 부동산 투자\1-1. 부동산 통계\☆ 1-1. KB 부동산 데이터 차트.xlsx"),
    Path(r"C:\Users\smk20\Desktop\☆ 1-1. KB 부동산 데이터 차트.xlsx"),
]

# ────────────────────────────────────────────────────────────────
# 매매지수/전세지수 시트: row2=한글헤더, row3=영문헤더, row4~=데이터
# 열 인덱스(0-based) 1~186 = 전국~기타지방
# (col_idx, region_code, region_name)
# ────────────────────────────────────────────────────────────────
ALL_PRICE_REGIONS = [
    (1,   "0000",   "전국"),
    (2,   "1100",   "서울"),
    (3,   "1101",   "서울 강북"),
    (4,   "KB004",  "강북구"),
    (5,   "KB005",  "광진구"),
    (6,   "KB006",  "노원구"),
    (7,   "KB007",  "도봉구"),
    (8,   "KB008",  "동대문구"),
    (9,   "KB009",  "마포구"),
    (10,  "KB010",  "서대문구"),
    (11,  "KB011",  "성동구"),
    (12,  "KB012",  "성북구"),
    (13,  "KB013",  "용산구"),
    (14,  "KB014",  "은평구"),
    (15,  "KB015",  "종로구"),
    (16,  "KB016",  "서울 중구"),
    (17,  "KB017",  "중랑구"),
    (18,  "1102",   "서울 강남"),
    (19,  "KB019",  "강남구"),
    (20,  "KB020",  "강동구"),
    (21,  "KB021",  "강서구"),
    (22,  "KB022",  "관악구"),
    (23,  "KB023",  "구로구"),
    (24,  "KB024",  "금천구"),
    (25,  "KB025",  "동작구"),
    (26,  "KB026",  "서초구"),
    (27,  "KB027",  "송파구"),
    (28,  "KB028",  "양천구"),
    (29,  "KB029",  "영등포구"),
    (30,  "KB030",  "6개광역시"),
    (31,  "2600",   "부산"),
    (32,  "KB032",  "부산 중구"),
    (33,  "KB033",  "부산 서구"),
    (34,  "KB034",  "부산 동구"),
    (35,  "KB035",  "부산 영도구"),
    (36,  "KB036",  "부산진구"),
    (37,  "KB037",  "부산 동래구"),
    (38,  "KB038",  "부산 남구"),
    (39,  "KB039",  "부산 북구"),
    (40,  "KB040",  "부산 해운대구"),
    (41,  "KB041",  "부산 사하구"),
    (42,  "KB042",  "부산 금정구"),
    (43,  "KB043",  "부산 연제구"),
    (44,  "KB044",  "부산 수영구"),
    (45,  "KB045",  "부산 사상구"),
    (46,  "KB046",  "부산 기장군"),
    (47,  "KB047",  "부산 강서구"),
    (48,  "2700",   "대구"),
    (49,  "KB049",  "대구 중구"),
    (50,  "KB050",  "대구 동구"),
    (51,  "KB051",  "대구 서구"),
    (52,  "KB052",  "대구 남구"),
    (53,  "KB053",  "대구 북구"),
    (54,  "KB054",  "대구 수성구"),
    (55,  "KB055",  "대구 달서구"),
    (56,  "KB056",  "대구 달성군"),
    (57,  "2300",   "인천"),
    (58,  "KB058",  "인천 중구"),
    (59,  "KB059",  "인천 동구"),
    (60,  "KB060",  "인천 미추홀구"),
    (61,  "KB061",  "인천 연수구"),
    (62,  "KB062",  "인천 남동구"),
    (63,  "KB063",  "인천 부평구"),
    (64,  "KB064",  "인천 계양구"),
    (65,  "KB065",  "인천 서구"),
    (66,  "2900",   "광주"),
    (67,  "KB067",  "광주 동구"),
    (68,  "KB068",  "광주 서구"),
    (69,  "KB069",  "광주 남구"),
    (70,  "KB070",  "광주 북구"),
    (71,  "KB071",  "광주 광산구"),
    (72,  "3000",   "대전"),
    (73,  "KB073",  "대전 동구"),
    (74,  "KB074",  "대전 중구"),
    (75,  "KB075",  "대전 서구"),
    (76,  "KB076",  "대전 유성구"),
    (77,  "KB077",  "대전 대덕구"),
    (78,  "3100",   "울산"),
    (79,  "KB079",  "울산 중구"),
    (80,  "KB080",  "울산 남구"),
    (81,  "KB081",  "울산 동구"),
    (82,  "KB082",  "울산 북구"),
    (83,  "KB083",  "울산 울주군"),
    (84,  "KB084",  "5개광역시"),
    (85,  "KB085",  "수도권"),
    (86,  "3600",   "세종"),
    (87,  "4100",   "경기"),
    (88,  "KB088",  "수원"),
    (89,  "KB089",  "수원 장안구"),
    (90,  "KB090",  "수원 권선구"),
    (91,  "KB091",  "수원 팔달구"),
    (92,  "KB092",  "수원 영통구"),
    (93,  "KB093",  "성남"),
    (94,  "KB094",  "성남 수정구"),
    (95,  "KB095",  "성남 중원구"),
    (96,  "KB096",  "성남 분당구"),
    (97,  "KB097",  "고양"),
    (98,  "KB098",  "고양 덕양구"),
    (99,  "KB099",  "일산동구"),
    (100, "KB100",  "일산서구"),
    (101, "KB101",  "안양"),
    (102, "KB102",  "안양 만안구"),
    (103, "KB103",  "안양 동안구"),
    (104, "KB104",  "부천"),
    (105, "KB105",  "의정부"),
    (106, "KB106",  "광명"),
    (107, "KB107",  "평택"),
    (108, "KB108",  "안산"),
    (109, "KB109",  "안산 단원구"),
    (110, "KB110",  "안산 상록구"),
    (111, "KB111",  "과천"),
    (112, "KB112",  "구리"),
    (113, "KB113",  "남양주"),
    (114, "KB114",  "용인"),
    (115, "KB115",  "용인 처인구"),
    (116, "KB116",  "용인 기흥구"),
    (117, "KB117",  "용인 수지구"),
    (118, "KB118",  "시흥"),
    (119, "KB119",  "군포"),
    (120, "KB120",  "의왕"),
    (121, "KB121",  "하남"),
    (122, "KB122",  "오산"),
    (123, "KB123",  "파주"),
    (124, "KB124",  "이천"),
    (125, "KB125",  "안성"),
    (126, "KB126",  "김포"),
    (127, "KB127",  "양주"),
    (128, "KB128",  "동두천"),
    (129, "KB129",  "광주(경기)"),
    (130, "KB130",  "화성"),
    (131, "4200",   "강원"),
    (132, "KB132",  "춘천"),
    (133, "KB133",  "강릉"),
    (134, "KB134",  "원주"),
    (135, "4300",   "충북"),
    (136, "KB136",  "청주"),
    (137, "KB137",  "청주 상당구"),
    (138, "KB138",  "청주 서원구"),
    (139, "KB139",  "청주 청원구"),
    (140, "KB140",  "청주 흥덕구"),
    (141, "KB141",  "충주"),
    (142, "KB142",  "제천"),
    (143, "4400",   "충남"),
    (144, "KB144",  "천안"),
    (145, "KB145",  "천안 동남구"),
    (146, "KB146",  "천안 서북구"),
    (147, "KB147",  "공주"),
    (148, "KB148",  "아산"),
    (149, "KB149",  "논산"),
    (150, "KB150",  "계룡"),
    (151, "KB151",  "당진"),
    (152, "KB152",  "서산"),
    (153, "4500",   "전북"),
    (154, "KB154",  "전주"),
    (155, "KB155",  "전주 완산구"),
    (156, "KB156",  "전주 덕진구"),
    (157, "KB157",  "익산"),
    (158, "KB158",  "군산"),
    (159, "4600",   "전남"),
    (160, "KB160",  "목포"),
    (161, "KB161",  "순천"),
    (162, "KB162",  "광양"),
    (163, "KB163",  "여수"),
    (164, "4700",   "경북"),
    (165, "KB165",  "포항"),
    (166, "KB166",  "포항 남구"),
    (167, "KB167",  "포항 북구"),
    (168, "KB168",  "구미"),
    (169, "KB169",  "경산"),
    (170, "KB170",  "안동"),
    (171, "KB171",  "김천"),
    (172, "4800",   "경남"),
    (173, "KB173",  "창원"),
    (174, "KB174",  "마산합포구"),
    (175, "KB175",  "마산회원구"),
    (176, "KB176",  "창원 성산구"),
    (177, "KB177",  "창원 의창구"),
    (178, "KB178",  "창원 진해구"),
    (179, "KB179",  "양산"),
    (180, "KB180",  "거제"),
    (181, "KB181",  "진주"),
    (182, "KB182",  "김해"),
    (183, "KB183",  "통영"),
    (184, "5000",   "제주도"),
    (185, "KB185",  "제주/서귀포"),
    (186, "KB186",  "기타지방"),
]

# ────────────────────────────────────────────────────────────────
# 매수매도심리 / 전세수급 시트: 3열 그룹, 3번째 열이 지수값
# (value_col_idx, region_code, region_name)
# ────────────────────────────────────────────────────────────────
ALL_SENTIMENT_REGIONS = [
    (3,  "0000",   "전국"),
    (6,  "1100",   "서울"),
    (9,  "1101",   "서울 강북"),
    (12, "1102",   "서울 강남"),
    (15, "KB030",  "6개광역시"),
    (18, "KB084",  "5개광역시"),
    (21, "2600",   "부산"),
    (24, "2700",   "대구"),
    (27, "2300",   "인천"),
    (30, "2900",   "광주"),
    (33, "3000",   "대전"),
    (36, "3100",   "울산"),
    (39, "3600",   "세종"),
    (42, "KB085",  "수도권"),
    (45, "4100",   "경기"),
    (48, "4200",   "강원"),
    (51, "4300",   "충북"),
    (54, "4400",   "충남"),
    (57, "4500",   "전북"),
    (60, "4600",   "전남"),
    (63, "4700",   "경북"),
    (66, "4800",   "경남"),
    (69, "KB186",  "기타지방"),
    (72, "5000",   "제주"),
]


def _find_file() -> Path | None:
    if KB_FILE and KB_FILE.exists():
        return KB_FILE
    for p in _SEARCH_PATHS:
        if p.exists():
            return p
    return None


def _to_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        from datetime import timedelta
        return (date(1899, 12, 30) + timedelta(days=int(v)))
    return None


def _parse_price_sheet(wb, sheet_name: str, price_type: str, change_sheet: str | None = None) -> list[dict]:
    """매매지수 또는 전세지수 시트를 읽어 kb_price_index rows 반환"""
    ws = wb[sheet_name]
    rows_raw = list(ws.iter_rows(values_only=True))

    # 변동률 시트 로드
    change_map: dict[tuple, float] = {}
    if change_sheet and change_sheet in wb.sheetnames:
        ws_c = wb[change_sheet]
        rows_c = list(ws_c.iter_rows(values_only=True))
        for row in rows_c[16:]:
            d = _to_date(row[0])
            if d is None:
                continue
            for col_idx, region_code, _ in ALL_PRICE_REGIONS:
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        change_map[(d, region_code)] = float(row[col_idx])
                    except (TypeError, ValueError):
                        pass

    cutoff = date(date.today().year - 5, 1, 1)
    results = []
    for row in rows_raw[3:]:  # row4 이후
        d = _to_date(row[0])
        if d is None or d < cutoff:
            continue
        for col_idx, region_code, region_name in ALL_PRICE_REGIONS:
            if col_idx >= len(row) or row[col_idx] is None:
                continue
            try:
                val = float(row[col_idx])
            except (TypeError, ValueError):
                continue
            results.append({
                "survey_date": d,
                "region_code": region_code,
                "region_name": region_name,
                "price_type": price_type,
                "index_value": val,
                "weekly_change": change_map.get((d, region_code)),
                "jeonse_ratio": None,  # run_all()에서 매매/전세 교차 계산
            })

    logger.info(f"KB {sheet_name}({price_type}): {len(results)}건 ({len(ALL_PRICE_REGIONS)}개 지역)")
    return results


def _parse_sentiment_sheet(wb) -> list[dict]:
    """3. 매수매도심리 + 4. 전세수급 시트를 읽어 kb_sentiment rows 반환"""
    ws_buy = wb["3. 매수매도심리"]
    ws_rent = wb["4. 전세수급"]

    rows_buy = list(ws_buy.iter_rows(values_only=True))
    rows_rent = list(ws_rent.iter_rows(values_only=True))

    # 매수매도심리 3년, 전세수급 5년
    cutoff_buy = date(date.today().year - 3, 1, 1)
    cutoff_rent = date(date.today().year - 5, 1, 1)

    combined: dict[tuple, dict] = {}

    for row in rows_buy[3:]:
        d = _to_date(row[0])
        if d is None or d < cutoff_buy:
            continue
        for col_idx, region_code, region_name in ALL_SENTIMENT_REGIONS:
            if col_idx >= len(row) or row[col_idx] is None:
                continue
            try:
                val = float(row[col_idx])
            except (TypeError, ValueError):
                continue
            key = (d, region_code)
            if key not in combined:
                combined[key] = {"region_name": region_name, "buyer_seller": None, "jeonse_supply": None}
            combined[key]["buyer_seller"] = val

    for row in rows_rent[3:]:
        d = _to_date(row[0])
        if d is None or d < cutoff_rent:
            continue
        for col_idx, region_code, region_name in ALL_SENTIMENT_REGIONS:
            if col_idx >= len(row) or row[col_idx] is None:
                continue
            try:
                val = float(row[col_idx])
            except (TypeError, ValueError):
                continue
            key = (d, region_code)
            if key not in combined:
                combined[key] = {"region_name": region_name, "buyer_seller": None, "jeonse_supply": None}
            combined[key]["jeonse_supply"] = val

    results = [
        {
            "survey_date": d,
            "region_code": region_code,
            "region_name": v["region_name"],
            "buyer_seller_index": v["buyer_seller"],
            "jeonse_supply_index": v["jeonse_supply"],
        }
        for (d, region_code), v in combined.items()
    ]
    logger.info(f"KB 심리/수급: {len(results)}건 ({len(ALL_SENTIMENT_REGIONS)}개 지역)")
    return results


async def run_all() -> dict:
    path = _find_file()
    if path is None:
        logger.error("KB 엑셀 파일을 찾을 수 없습니다")
        return {"price": [], "sentiment": []}

    logger.info(f"KB 파일 로드: {path}")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"KB 파일 열기 실패: {e}")
        return {"price": [], "sentiment": []}

    price_buy = _parse_price_sheet(wb, "매매지수", "buy", "매매증감")
    price_rent = _parse_price_sheet(wb, "전세지수", "rent", "전세증감")
    sentiment = _parse_sentiment_sheet(wb)

    wb.close()

    # 전세가율 계산: (survey_date, region_code) 기준 rent.index / buy.index
    rent_lookup = {
        (r["survey_date"], r["region_code"]): r["index_value"]
        for r in price_rent
    }
    for row in price_buy:
        rent_val = rent_lookup.get((row["survey_date"], row["region_code"]))
        if rent_val is not None and row["index_value"] and row["index_value"] > 0:
            row["jeonse_ratio"] = rent_val / row["index_value"]

    return {
        "price": price_buy + price_rent,
        "sentiment": sentiment,
    }
